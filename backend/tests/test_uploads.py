from io import BytesIO

from openpyxl import Workbook

from app.services.platform import uploads
from app.services.platform.uploads import UploadService


def test_save_csv_returns_file_id_and_summary(tmp_path):
    service = UploadService(upload_dir=tmp_path)
    content = b"quarter,revenue\nQ1,100\nQ2,125\n"

    result = service.save_upload(filename="revenue.csv", content=content)

    assert result.file_id.endswith(".csv")
    assert result.filename == "revenue.csv"
    assert result.row_count == 2
    assert result.columns == ["quarter", "revenue"]
    assert "Q1" in result.preview
    assert service.get_summary(result.file_id).row_count == 2


def test_save_header_only_csv_preserves_columns(tmp_path):
    service = UploadService(upload_dir=tmp_path)

    result = service.save_upload(filename="empty.csv", content=b"quarter,revenue\n")

    assert result.row_count == 0
    assert result.columns == ["quarter", "revenue"]


def test_save_xlsx_returns_file_id_and_summary(tmp_path):
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["quarter", "revenue"])
    sheet.append(["Q1", 100])
    sheet.append(["Q2", 125])
    buffer = BytesIO()
    workbook.save(buffer)

    service = UploadService(upload_dir=tmp_path)
    result = service.save_upload(filename="revenue.xlsx", content=buffer.getvalue())

    assert result.file_id.endswith(".xlsx")
    assert result.filename == "revenue.xlsx"
    assert result.row_count == 2
    assert result.columns == ["quarter", "revenue"]
    assert "Q1" in result.preview


def test_get_rows_returns_xlsx_rows(tmp_path):
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["quarter", "revenue"])
    sheet.append(["Q1", 100])
    sheet.append(["Q2", 125])
    buffer = BytesIO()
    workbook.save(buffer)

    service = UploadService(upload_dir=tmp_path)
    result = service.save_upload(filename="revenue.xlsx", content=buffer.getvalue())

    assert service.get_rows(result.file_id) == [
        {"quarter": "Q1", "revenue": "100"},
        {"quarter": "Q2", "revenue": "125"},
    ]


def test_save_xlsx_does_not_materialize_worksheet_rows(tmp_path, monkeypatch):
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["quarter", "revenue"])
    sheet.append(["Q1", 100])
    buffer = BytesIO()
    workbook.save(buffer)

    def fail_on_list(_value):
        raise AssertionError("XLSX parser should stream worksheet rows")

    monkeypatch.setattr(uploads, "list", fail_on_list, raising=False)

    service = UploadService(upload_dir=tmp_path)
    result = service.save_upload(filename="revenue.xlsx", content=buffer.getvalue())

    assert result.row_count == 1


def test_save_header_only_xlsx_preserves_columns(tmp_path):
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["quarter", "revenue"])
    buffer = BytesIO()
    workbook.save(buffer)

    service = UploadService(upload_dir=tmp_path)
    result = service.save_upload(filename="empty.xlsx", content=buffer.getvalue())

    assert result.row_count == 0
    assert result.columns == ["quarter", "revenue"]
