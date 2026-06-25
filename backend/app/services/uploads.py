import csv
import io
import json
import time
import uuid
from pathlib import Path

from openpyxl import load_workbook

from app.config import settings
from app.models.schemas import UploadResponse


class UploadService:
    def __init__(self, upload_dir: str | Path | None = None):
        self.upload_dir = Path(upload_dir or settings.local_upload_dir)
        self.upload_dir.mkdir(parents=True, exist_ok=True)

    def save_upload(self, *, filename: str, content: bytes) -> UploadResponse:
        suffix = Path(filename).suffix.lower()
        if suffix not in settings.allowed_upload_extensions:
            raise ValueError(f"Unsupported file type: {suffix}")
        if len(content) > settings.max_upload_bytes:
            raise ValueError("Upload exceeds maximum size")

        file_id = f"{uuid.uuid4()}{suffix}"
        response = self._summarize(file_id=file_id, filename=filename, content=content)
        path = self.upload_dir / file_id
        path.write_bytes(content)
        self._metadata_path(file_id).write_text(json.dumps({"filename": filename}), encoding="utf-8")
        return response

    def get_summary(self, file_id: str) -> UploadResponse:
        if Path(file_id).name != file_id:
            raise ValueError("Invalid file id")
        path = self.upload_dir / file_id
        if not path.exists():
            raise FileNotFoundError(file_id)
        return self._summarize(file_id=file_id, filename=self._original_filename(file_id), content=path.read_bytes())

    def get_rows(self, file_id: str) -> list[dict[str, str]]:
        if Path(file_id).name != file_id:
            raise ValueError("Invalid file id")
        path = self.upload_dir / file_id
        if not path.exists():
            raise FileNotFoundError(file_id)

        content = path.read_bytes()
        suffix = Path(file_id).suffix.lower()
        if suffix == ".csv":
            return self._rows_from_csv(content)
        if suffix == ".xlsx":
            return self._rows_from_xlsx(content)
        raise ValueError(f"Unsupported file type: {suffix}")

    def get_ai_summary(self, file_id: str) -> dict[str, object]:
        summary = self.get_summary(file_id)
        return {
            "filename": summary.filename,
            "columns": summary.columns,
            "row_count": summary.row_count,
            "preview": summary.preview,
        }

    def purge_expired(self, max_age_seconds: int) -> int:
        now = time.time()
        count = 0
        for path in self.upload_dir.glob("*"):
            if (
                path.is_file()
                and path.suffix.lower() in settings.allowed_upload_extensions
                and now - path.stat().st_mtime > max_age_seconds
            ):
                path.unlink()
                metadata_path = self._metadata_path(path.name)
                if metadata_path.exists():
                    metadata_path.unlink()
                count += 1
        return count

    def _metadata_path(self, file_id: str) -> Path:
        return self.upload_dir / f"{file_id}.json"

    def _original_filename(self, file_id: str) -> str:
        metadata_path = self._metadata_path(file_id)
        if not metadata_path.exists():
            return file_id
        try:
            data = json.loads(metadata_path.read_text(encoding="utf-8"))
        except json.JSONDecodeError:
            return file_id
        filename = data.get("filename")
        return filename if isinstance(filename, str) and filename else file_id

    def _summarize(self, *, file_id: str, filename: str, content: bytes) -> UploadResponse:
        suffix = Path(file_id).suffix.lower()
        if suffix == ".csv":
            text = content.decode("utf-8-sig")
            reader = csv.DictReader(io.StringIO(text))
            rows = list(reader)
            columns = reader.fieldnames or []
            preview = "\n".join(text.splitlines()[:4])
            return UploadResponse(
                file_id=file_id,
                filename=filename,
                row_count=len(rows),
                columns=columns,
                preview=preview,
            )
        if suffix == ".xlsx":
            columns, rows = self._parse_xlsx(content)
            preview_rows = [",".join(columns)]
            preview_rows.extend(",".join(row[column] for column in columns) for row in rows[:3])
            return UploadResponse(
                file_id=file_id,
                filename=filename,
                row_count=len(rows),
                columns=columns,
                preview="\n".join(preview_rows),
            )
        raise ValueError(f"Unsupported file type: {suffix}")

    def _rows_from_csv(self, content: bytes) -> list[dict[str, str]]:
        text = content.decode("utf-8-sig")
        return list(csv.DictReader(io.StringIO(text)))

    def _rows_from_xlsx(self, content: bytes) -> list[dict[str, str]]:
        return self._parse_xlsx(content)[1]

    def _parse_xlsx(self, content: bytes) -> tuple[list[str], list[dict[str, str]]]:
        workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        try:
            sheet = workbook.active
            row_values = sheet.iter_rows(values_only=True)
            header = next(row_values, None)
            if header is None:
                return [], []

            columns = ["" if value is None else str(value) for value in header]
            rows: list[dict[str, str]] = []
            for values in row_values:
                row = {
                    column: "" if value is None else str(value)
                    for column, value in zip(columns, values, strict=False)
                }
                rows.append(row)
            return columns, rows
        finally:
            workbook.close()
